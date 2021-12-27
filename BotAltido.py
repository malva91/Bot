import gspread
from oauth2client.service_account import ServiceAccountCredentials
from pprint import pprint
from openpyxl import load_workbook
import pickle
import datetime
from datetime import datetime
import telebot
from telebot import types
import os
import xlsxwriter

#variabili API google
scope =  ["https://spreadsheets.google.com/feeds",'https://www.googleapis.com/auth/spreadsheets',"https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"]
creds = ServiceAccountCredentials.from_json_keyfile_name("infoaltidobot-1415ef45e65f.json", scope)
client = gspread.authorize(creds)
sheet = client.open_by_key('1bGhvpZco9fRKX56SNdu7VwHkOjZppBEAt5SFZAyCauo').sheet1
#variabili API telegram
API = sheet.acell('A2').value
API_TOKEN = API
tb = telebot.TeleBot(API_TOKEN, parse_mode='HTML')
#variabili globali
cartella=load_workbook(filename='Variabili.xlsx', data_only=True)
print (cartella.sheetnames)




bot = telebot.TeleBot(API_TOKEN)


# Variabili appartamenti
spazio = '\n\n\n--------------------------------------\n\n\n'
bovio = 'Indirizzo:\nViao Giovanni Bovio, 9,Pisa\n\nNome:\nVia Bovio\n\nCodice:\nPIPSIF546\n\nComposizione:\n1 Matrimoniale\n1 Bagno\n1 Cucina\n\nNote:\nNOI ABBIAMO CHIAVE DEL PTONE QUELLA DELL APPARTAMENTO è NEL PTA OMBRELLO'

girasole = 'Indirizzo:\nVia delle Selvette, 233, 55018, Capannori, Lucca\n\nNome:\nGirasole\n\nCodice:\nLUSIMF1081\n\nComposizione:\n1 Matrimoniale\n2 Singoli\n1 Bagno\n1 Cucina\n\nNote:\n'
hibiscus = 'Indirizzo:\nVia delle Selvette, 233, 55018, Capannori, Lucca\n\nNome:\nHibiscus\n\nCodice:\nLUSIMF1082\n\nComposizione:\n1 Matrimoniale\n2 Singoli (fatti a divano in cucina)\n1 Bagno\n1 Cucina\n\nNote:\n'
iris = 'Indirizzo:\nVia delle Selvette, 233, 55018, Capannori, Lucca\n\nNome:\nIris\n\nCodice:\nLUSIMF1083\n\nComposizione:\n1 Matrimoniale KING\n3 Singoli\n2 Bagno\n1 Cucina\n\nNote:\n'
orchidea = 'Indirizzo:\nVia delle Selvette, 233, 55018, Capannori, Lucca\n\nNome:\nOrchidea\n\nCodice:\nLUSIMF1084\n\nComposizione:\n1 Matrimoniale KING\n2 Singoli\n2 Bagno\n1 Cucina\n\nNote:\n'
rosmarino = 'Indirizzo:\nVia delle Selvette, 233, 55018, Capannori, Lucca\n\nNome:\nRosmarino\n\nCodice:\nLUSIMF1085\n\nComposizione:\n3 Matrimoniale\n2 Bagno\n1 Cucina\n\nNote:\n'
rosa =  'Indirizzo:\nVia delle Selvette, 233, 55018, Capannori, Lucca\n\nNome:\nRosa\n\nCodice:\nLUSIMF1086\n\nComposizione:\n1 Matrimoniale\n2 Singoli \n2 Bagno\n1 Cucina\n\nNote:\n'
segrohimigno = girasole+spazio+hibiscus+spazio+iris+spazio+orchidea+spazio+rosmarino+spazio+rosa


pisano = 'Indirizzo:\nViale Giovanni Pisano, 10, Pisa\n\nNome:\nVia Pisano\n\nCodice:\nPIPSIF464\n\nComposizione:\n1 Matrimoniale\n1 Bagno\n1 Cucina\n\nNote:\n'
puccini = 'Indirizzo:\nVia Giacomo Puccini, 151, Viareggio\n\nNome:\nVia Puccini\n\nCodice:\nLUVRGF747\n\nComposizione:\n2 Matrimoniale\n2 Divano Letto\n2 Singoli\n2 Bagno\n1 Cucina\n\nNote:\nABBIAMO LA CHIAVE DEL PORTONE MENTRE DELLA CASA è NEL PORTA OMBRELLO'
ville = 'Indirizzo:\n/\n\nNome:\nVia delle Ville\n\nCodice:\nLUMSGF997\n\nComposizione:\n\nDati ancora mancanti\n\nNote:\n'
alessio ='Indirizzo:\nVia di Sant Alessio, 25, Lucca\n\nNome:\nVia di Sant Alessio\n\nCodice:\nLULCCF433\n\nComposizione:\n2 Matrimoniale\n1 Bagno\n1Cucina\n\nNote:\n'
antonio = 'Indirizzo:\nVia di Sant Antonio, 8, Pisa\n\nNome:\nVia di Sant Antonio\n\nCodice:\nPIPSIF1102\n\nComposizione:\n1 Matrimoniale\n1 Divano letto\n1 Bagno\n1 Cucina\n\nNote:\n'
mancini = 'Indirizzo:\nVia Augusto Mancini, 49, Lucca\n\nNome:\nVia Mancini\n\nCodice:\nLULCCF686\n\nComposizione:\n\nDati ancora mancanti\n\nNote:\nLE CHIAVI SONO NELLA CASSETTA DELLA POSTA'
facchini = 'Indirizzo:\nVia dei Facchini, 10, Pisa\n\nNome:\nVia Facchini\n\nCodice:\nPIPSIF635\n\nComposizione:\n1 Matrimoniale\n1(2) Sigoli a castello\n1 Bagno\n1 Cucina\n\nNote:\n'
vico = 'Indirizzo:\nVia dei Facchini, 10, Pisa\n\nNome:\nVia Facchini\n\nCodice:\nLUFRTF468\n\nComposizione:\n2 Matrimoniale\n1 Sigolo\n1 Poltrola Letto\n1 Bagno\n1 Cucina\n\nNote:\n'
luigiPera = 'Indirizzo:\nVia Luigi Pera ,14, Pisa\n\nNome:\nCisanello\n\nCodice:\nPIPSIF1106\n\nComposizione:\n1 Matrimoniale\n1 Bagno\n1 Cucina\nNote:\n'

gagnoVerde = 'Indirizzo:\nVia di Gagno, 50, Pisa (Primo piano)\n\nNome:\nGagno verde\n\nCodice:\nPIPSIF720\n\nComposizione:\n1 Matrimoniale\n1 Divano Letto\n1 Bagno\n1 Cucina\n\nNote:\n'
gagnoBlu = 'Indirizzo:\nVia di Gagno, 50, Pisa (Secondo piano)\n\nNome:\nGagno blu\n\nCodice:\nPIPSIF719\n\nComposizione:\n1 Matrimoniale\n1 Divano Letto\n1 Bagno\n1 Cucina\n\nNote:\n'
gagnoGiallo = 'Indirizzo:\nVia di Gagno, 50, Pisa (Penultimo piano)\n\nNome:\nGagno giallo\n\nCodice:\nPIPSIF1096\n\nComposizione:\n1 Matrimoniale\n1 Divano Letto\n1 Bagno\n1 Cucina\n\nNote:\n'
gagnoRosso = 'Indirizzo:\nVia di Gagno, 50, Pisa (Ultimo piano)\n\nNome:\nGagno rosso\n\nCodice:\nPIPSIF721\n\nComposizione:\n2 Matrimoniale\n1 Divano Letto\n1 Bagno\n1 Cucina\n\nNote:\n'
gagno = gagnoVerde+spazio+gagnoBlu+spazio+gagnoGiallo+spazio+gagnoGiallo

lenzeTot = 'Indirizzo:\nVia delle Lenze, 122, 56122 Pisa\n\nNome:\nVia delle Lenze TOTALE\n\nCodice:\nPIPSIF1031\n\nComposizione:\n2 Matrimoniale\n2 Singoli\n3 Bagni\n1 Cucina\n\nNote:\n'
lenzePT = 'Indirizzo:\nVia delle Lenze, 122, 56122 Pisa\n\nNome:\nVia delle Lenze PIANO TERRA\n\nCodice:\nPIPSIF1032\n\nComposizione:\n1 Matrimoniale\n1 Bagno\nNote:\n'
lenze1P = 'Indirizzo:\nVia delle Lenze, 122, 56122 Pisa\n\nNome:\nVia delle Lenze PRIMO PIANO\n\nCodice:\nPIPSIF1033\n\nComposizione:\n2 Singoli\n1 Bagno\nNote:\n'
lenze1PT = 'Indirizzo:\nVia delle Lenze, 122, 56122 Pisa\n\nNome:\nVia delle Lenze PRIMO PIANO CON TERRAZZO\n\nCodice:\nPIPSIF10334\n\nComposizione:\n1 Matrimoniale\n1 Bagno\nNote:\n'
lenze = lenzePT+spazio+lenze1P+spazio+lenze1PT







lista = '-Girasole\n-Hibiscus\n-iris\n-orchidea\n-Rosmarino\n-Rosa\n-Segromigno(gli raggruppa)\n\n-Gagno rosso\nGagno giallo\n-Gagno verde\n-Gagno blu\n-Gagno(gli raggruppa)\n\n-Lenze totale\n-Lenze pt(Piano terra)\n-Lenze 1p(primo piano\n-Lenze 1pt(primo piano con terrazzo)\n-Lenze (gli raggruppa)\n\n-Pisano\n-Puccini\n-Ville (Via delle Ville)\n-Alessio (Sant Alessio)\nAntonio (Sant Antonio)\n-Mancini\n-Bovio\n-Facchini\n-Vico\n-LuigiPera(va bene anche solo Luigi o Pera)'
 
aiuto = '- Puoi inserire nome o codice dell appartamento\n\n- "Lista" avrai la lista completa degli appartamenti\n\n- "Segromigno" raggruppa gli appartaemnti di Segromigno\n\n- "Gagno" raggruppa gli appartamenti di gagno\n\n- "Lenze" raggruppa gli appartaemnti di Via delle Lenze'


# Handle '/start' and '/help'
@bot.message_handler(commands=['help','start'])
def send_welcome(message):
    bot.reply_to(
        message, 'Ciao, di che appartamento hai bisogno?')


# Handle all other messages with content_type 'text' (content_types defaults to ['text'])
@bot.message_handler(func=lambda message: True)
def echo_message(message):
    
    if message.text == 'Aiuto' or message.text == 'aiuto':
        bot.reply_to(message, aiuto)

    elif message.text == 'Bovio' or message.text == 'bovio' or message.text == 'PIPSIF546' or message.text == '546':
        bot.reply_to(message,bovio)
        
    elif message.text == 'Segromigno' or message.text == 'segromigno':
        bot.reply_to(message, segrohimigno)

    elif message.text == 'Girasole' or message.text == 'girasole' or message.text == 'LUSIMF1081' or message.text == '1081':
        bot.reply_to(message, girasole)

    elif message.text == 'Hibiscus' or message.text == 'hibiscus' or message.text == 'LUSIMF1082' or message.text == '1082':
        bot.reply_to(message, hibiscus )

    elif message.text == 'Iris' or message.text == 'iris' or message.text == 'LUSIMF1083' or message.text == '1083':
        bot.reply_to(message, iris )

    elif message.text == 'Orchidea' or message.text == 'orchidea' or message.text == 'LUSIMF1084' or message.text == '1084':
        bot.reply_to(message, orchidea )

    elif message.text == 'Rosmarino' or message.text == 'rosmarino' or message.text == 'LUSIMF1085' or message.text == '1085':
        bot.reply_to(message, rosmarino )

    elif message.text == 'Rosa' or message.text == 'rosa' or message.text == 'LUSIMF1086' or message.text == '1086':
        bot.reply_to(message, rosa)

    elif message.text == 'Pisano' or message.text == 'pisano' or message.text == 'PIPSIF464' or message.text == '464':
        bot.reply_to(message, pisano)

    elif message.text == 'Puccini' or message.text == 'puccini' or message.text == 'LUVRGF747' or message.text == '747':
        bot.reply_to(message, puccini)

    elif message.text == 'Ville' or message.text == 'ville' or message.text == 'LUMSGF997' or message.text == '997':
        bot.reply_to(message, ville)

    elif message.text == 'Alessio' or  message.text == 'alessio' or message.text == 'LULCCF433' or message.text == '433':
        bot.reply_to(message, alessio)

    elif message.text == 'Antonio' or  message.text == 'antonio' or message.text == 'PIPSIF1102' or message.text == '1102':
        bot.reply_to(message, antonio)


    elif message.text == 'Mancini' or  message.text == 'mancini' or message.text == 'LULCCF686' or message.text == '686':
        bot.reply_to(message, mancini)

    elif message.text == 'Gagno Green' or  message.text == 'gagno green' or message.text == 'Gagno green' or message.text == 'gagno Green'or  message.text == 'Gagno Verde' or message.text == 'gagno verde' or message.text == 'Gagno verde' or message.text == 'gagno Verde' or message.text == 'PIPSIF720' or message.text == '720':
        bot.reply_to(message, gagnoVerde)

    elif message.text == 'Gagno Blu' or  message.text == 'gagno blu' or message.text == 'Gagno blu' or message.text == 'gagno Blu'or message.text == 'PIPSIF719' or message.text == '719':
        bot.reply_to(message, gagnoBlu)

    elif message.text == 'Gagno Yellow' or  message.text == 'gagno yellow' or message.text == 'Gagno yellow' or message.text == 'gagno Yellow'or  message.text == 'Gagno Giallo' or message.text == 'gagno giallo' or message.text == 'Gagno giallo' or message.text == 'gagno Giallo' or message.text == 'PIPSIF1096' or message.text == '1096':
        bot.reply_to(message, gagnoGiallo)

    elif message.text == 'Gagno Red' or  message.text == 'gagno red' or message.text == 'Gagno red' or message.text == 'gagno Red'or  message.text == 'Gagno Rosso' or message.text == 'gagno rosso' or message.text == 'Gagno rosso' or message.text == 'gagno Rosso' or message.text == 'PIPSIF721' or message.text == '721':
        bot.reply_to(message, gagnoRosso)


    elif message.text == 'Gagno' or  message.text == 'gagno':
        bot.reply_to(message, gagno)


    elif message.text == 'Lenze totale' or  message.text == 'lenze totale' or message.text == 'PIPSIF1031' or message.text == '1031':
        bot.reply_to(message, lenzeTot)

    elif message.text == 'Lenze pt' or message.text == 'lenze pt' or message.text == 'PIPSIF1032' or message.text == '1032':
        bot.reply_to(message, lenzePT)

    elif message.text == 'Lenze 1p' or message.text == 'lenze 1p' or message.text == 'PIPSIF1033' or message.text == '1033':
        bot.reply_to(message, lenze1P)

    elif message.text == 'Lenze 1pt' or message.text == 'lenze 1pt' or message.text == 'PIPSIF1034' or message.text == '1034':
        bot.reply_to(message, lenze1PT)

    elif message.text == 'Lenze' or message.text == 'lenze':
        bot.reply_to(message,lenze)

    elif message.text == 'Fachini' or message.text == 'facchini' or message.text == 'PIPSIF635' or message.text == '635':
        bot.reply_to(message, facchini )

    elif message.text == 'Vico' or message.text == 'vico' or message.text == 'LUFRTF468' or message.text == '468':
        bot.reply_to(message, vico)

    elif message.text == 'pera' or message.text == 'Pera'or message.text == 'luigi' or message.text == 'Luigi'or message.text == '1106' or message.text == 'PIPSIF1106':
        bot.reply_to(message, luigiPera)


    elif message.text == 'Lista' or message.text == 'lista':
        bot.reply_to(message, lista)




    else:
        bot.reply_to(message, 'Non riconosco l appartamento')


bot.infinity_polling()
